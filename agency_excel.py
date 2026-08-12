# -*- coding: utf-8 -*-
"""
代理商 CUE Excel 渲染模組 (Agency Cue Excel Renderer)

以 openpyxl 精緻復刻三種代理商版型（2008傳媒 / 佳聖 / 凱絡），每平台一個工作表。
字級、列高、框線層次、底色、格式代碼、圖片錨點、頁首頁尾字串皆逐格對齊原始範例檔實測值。

- 字型一律「微軟正黑體」；版面靠 fitToWidth 縮放，不靠縮小字級塞版。
- 金額：2008 與凱絡三層價用會計 $ 格式；佳聖 金額與凱絡費用區用無 $ 純數字。
- 為求 LibreOffice/Excel 轉 PDF 穩定，數字一律寫入計算後實值（非公式）。
"""
import os
import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

import config
import agency_cue as ac

FONT = config.FONT_MAIN

# 數字格式
ACCT = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'   # 會計 $（2008、凱絡三層價）
NUM = "#,##0_);[Red](#,##0)"          # 無 $ 純數字（佳聖 金額、凱絡費用）
DAY_FMT = "0_ "                        # 每日檔次（尾隨空格）
DAY_RED = "#,##0_);[Red](#,##0)"      # 2008 每日合計
SUM_FMT = "#,##0_);(#,##0)"           # 2008 合計 F
MATERIAL_FMT = 'm"月"d"日"'            # 佳聖 素材日期
CARAT_H_FMT = "0_);[Red]\\(0\\)"      # 凱絡 檔數
CN_WD = "一二三四五六日"
EN_WD = "MTWTFSS"

YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")   # 週末亮黃


# =============================================================================
# 共用工具
# =============================================================================
def _cn_weekday(d):
    return CN_WD[d.weekday()]


def _en_weekday(d):
    return EN_WD[d.weekday()]


def _set(ws, coord, value, size=12, bold=False, align="center", valign="center",
         wrap=False, fmt=None, fill=None):
    c = ws[coord]
    c.value = value
    c.font = Font(name=FONT, size=size, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def _merge(ws, rng):
    ws.merge_cells(rng)


def _edge(ws, coord, top=None, bottom=None, left=None, right=None):
    """疊加單元格邊框（不覆蓋既有其他邊）。"""
    c = ws[coord]
    b = c.border
    c.border = Border(
        left=Side(style=left) if left else b.left,
        right=Side(style=right) if right else b.right,
        top=Side(style=top) if top else b.top,
        bottom=Side(style=bottom) if bottom else b.bottom,
    )


def _box(ws, r1, c1, r2, c2, edge="medium", inner="thin"):
    """畫外框 edge、內線 inner 的方框（inner=None 則只有外框）。"""
    es = Side(style=edge)
    ins = Side(style=inner) if inner else Side(style=None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=es if c == c1 else ins,
                right=es if c == c2 else ins,
                top=es if r == r1 else ins,
                bottom=es if r == r2 else ins,
            )


def _net_cell(ws, coord, net_display, size, bold=False, fmt=ACCT, align="center", fill=None):
    """實收欄：數字用指定金額格式，字串（專案回饋/計價於量販）用文字。"""
    if isinstance(net_display, str):
        _set(ws, coord, net_display, size=size, bold=bold, align=align, fill=fill)
    else:
        _set(ws, coord, net_display, size=size, bold=bold, fmt=fmt, align=align, fill=fill)


def _page(ws, orientation="landscape", margins=(0.31, 0.31, 1.1, 0.75)):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    l, r, t, b = margins
    ws.page_margins = PageMargins(left=l, right=r, top=t, bottom=b)


def _write_day_header(ws, start_col, start_dt, days, r_month, r_date, r_wd,
                      weekday_fn, month_style="en", size=12, date_as_datetime=False,
                      date_fmt="d", weekend_fill_rows=()):
    """
    寫日期表頭三列：月份(同月合併)、日號、星期。
    weekend_fill_rows: 需在週六日填黃底的列號集合。
    """
    from datetime import timedelta
    # 月份列（同月合併）
    seg_start = 0
    for i in range(1, days + 1):
        cur = start_dt + timedelta(days=i) if i < days else None
        prev = start_dt + timedelta(days=i - 1)
        boundary = (i == days) or (cur.month != prev.month)
        if boundary:
            c0 = start_col + seg_start
            c1 = start_col + i - 1
            mon = (start_dt + timedelta(days=seg_start)).month
            label = calendar.month_abbr[mon] if month_style == "en" else f"{mon}月"
            _set(ws, f"{get_column_letter(c0)}{r_month}", label, size=size, bold=(month_style == "en"),
                 align="left" if month_style == "en" else "center")
            if c1 > c0:
                _merge(ws, f"{get_column_letter(c0)}{r_month}:{get_column_letter(c1)}{r_month}")
            seg_start = i
    # 日號 + 星期
    for i in range(days):
        d = start_dt + timedelta(days=i)
        col = get_column_letter(start_col + i)
        wend = d.weekday() >= 5
        f_date = YELLOW if (wend and r_date in weekend_fill_rows) else None
        f_wd = YELLOW if (wend and r_wd in weekend_fill_rows) else None
        if date_as_datetime:
            _set(ws, f"{col}{r_date}", d, size=size, bold=True, fmt=date_fmt, fill=f_date)
        else:
            _set(ws, f"{col}{r_date}", d.day, size=size, bold=True, fmt=date_fmt, fill=f_date)
        _set(ws, f"{col}{r_wd}", weekday_fn(d), size=size, fill=f_wd)


def _period_dot(a, b):
    return f"{a.strftime('%Y.%m.%d')}-{b.strftime('%Y.%m.%d')}"


def _period_short(a, b):
    return f"{a.month}/{a.day}-{b.month}/{b.day}"


def _mat_str(d):
    return f"{d.month}/{d.day}" if d else ""


# =============================================================================
# 代理商 Logo（真品，隨專案打包）
# =============================================================================
_ASSET_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
LOGO_2008 = os.path.join(_ASSET_DIR, "logo_2008.png")     # 150×77
LOGO_DDRIVE = os.path.join(_ASSET_DIR, "logo_ddrive.png")  # 145×176


def _add_logo(ws, path, anchor_cell, width, height):
    if not os.path.exists(path):
        return
    img = XLImage(path)
    img.width = width
    img.height = height
    img.anchor = anchor_cell
    ws.add_image(img)


def _group_main_comp(rows):
    """main(+緊接 comp) 為一組；其餘各自一組。"""
    groups = []
    i = 0
    while i < len(rows):
        r = rows[i]
        if r["kind"] == ac.KIND_MAIN and i + 1 < len(rows) and rows[i + 1]["kind"] == ac.KIND_COMP:
            groups.append([r, rows[i + 1]])
            i += 2
        else:
            groups.append([r])
            i += 1
    return groups


# =============================================================================
# 2008傳媒版型
# =============================================================================
def _render_2008(wb, sheet, model, made_date):
    days = (model["end_date"] - model["start_date"]).days + 1
    start_dt = model["start_date"]
    sec = sheet["seconds"]
    ws = wb.create_sheet(title=_sheet_name(model, sheet))
    _page(ws, margins=(0.31, 0.31, 1.1, 0.75))

    DAY0 = 9
    widths = {"A": 48.6, "B": 22.0, "C": 31.6, "D": 33.6, "E": 20.9, "F": 32.6, "G": 36.6, "H": 29.4}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for i in range(days):
        ws.column_dimensions[get_column_letter(DAY0 + i)].width = 8.9
    last_col = DAY0 + days - 1
    # 最後一欄（含跨月的 Sep）稍加寬，避免右側外框線被裁切遮住
    ws.column_dimensions[get_column_letter(last_col)].width = 11.5

    for r in range(1, 8):
        ws.row_dimensions[r].height = 40
    for r in (8, 9, 10):
        ws.row_dimensions[r].height = 41.9

    # Logo 右上（8.93cm × 4.85cm，96DPI 換算：寬 338px、高 183px）
    _add_logo(ws, LOGO_2008, f"{get_column_letter(max(DAY0, last_col - 5))}4", width=338, height=183)

    # 表頭 1~6：標籤靠左、值靠左
    header_pairs = [
        ("Client", model["client_name"]),
        ("Media Agency", "2008傳媒行銷股份有限公司"),
        ("Advertising Agency", ""),
        ("Product", model["product_name"]),
        ("Campaign", model.get("campaign", "")),
        ("Period", _period_dot(model["start_date"], model["end_date"])),
    ]
    for idx, (lab, val) in enumerate(header_pairs):
        rr = idx + 1
        _set(ws, f"A{rr}", lab, size=28, bold=True, align="left")
        _set(ws, f"C{rr}", val, size=28, bold=True, align="left")
        _merge(ws, f"C{rr}:H{rr}")

    # 欄位表頭 8~10（24pt 不粗）
    heads = [("A", "媒體型態"), ("B", "地區"), ("C", "播出時段"), ("D", "定價"),
             ("E", "單位"), ("F", "次數"), ("G", "合計"), ("H", "素材\n提供時間")]
    for col, txt in heads:
        _set(ws, f"{col}8", txt, size=24, wrap=True)
        _merge(ws, f"{col}8:{col}10")
    _write_day_header(ws, DAY0, start_dt, days, 8, 9, 10, _cn_weekday, month_style="en",
                      size=24, date_as_datetime=True, date_fmt="d", weekend_fill_rows=(9, 10))

    day_cols = [(DAY0 + i, i) for i in range(days)]

    # 資料列
    r = 11
    groups = _group_main_comp(sheet["rows"])
    data_top = r
    schedule_rows = []
    data_rows = []  # [(row_number, row_dict)]：供左側欄延伸合併判斷
    total_spots = total_list = total_net = 0
    for grp in groups:
        gtop, gbot = r, r + len(grp) - 1
        main = grp[0]
        for k, row in enumerate(grp):
            rr = r + k
            data_rows.append((rr, row))
            ws.row_dimensions[rr].height = 105.65 if row["kind"] == ac.KIND_MAIN else 93.75
            _set(ws, f"F{rr}", row["spots"], size=24, fmt=DAY_FMT)
            total_spots += row["spots"]
            if row["schedule"] is None:
                c0, c1 = get_column_letter(DAY0), get_column_letter(last_col)
                _set(ws, f"{c0}{rr}", row["spots"], size=24)
                if last_col > DAY0:
                    _merge(ws, f"{c0}{rr}:{c1}{rr}")
            else:
                for cidx, off in day_cols:
                    _set(ws, f"{get_column_letter(cidx)}{rr}", row["schedule"][off], size=24, fmt=DAY_FMT)
                schedule_rows.append((rr, row["schedule"]))
            if isinstance(row["net_display"], (int, float)):
                total_net += row["net_display"]
        _set(ws, f"A{gtop}", main["media_label"], size=24, wrap=True)
        _set(ws, f"B{gtop}", main["region_label"], size=24)
        _set(ws, f"C{gtop}", main["daypart"], size=24)
        # 樂家康(超市)定價比照合計欄，顯示「計價於量販」而非 $0
        if main["kind"] in (ac.KIND_SUPER, ac.KIND_SUPER_REBATE):
            _set(ws, f"D{gtop}", ac.NET_ON_MAG, size=24)
        else:
            _set(ws, f"D{gtop}", main["list_total"], size=24, fmt=ACCT)
        _set(ws, f"E{gtop}", f"{sec}秒", size=24)
        _net_cell(ws, f"G{gtop}", main["net_display"], size=24)
        _set(ws, f"H{gtop}", _mat_str(main["material"]), size=24, wrap=True)
        total_list += main["list_total"]
        # 定價/單位/合計：僅在同組（主+補償）內合併
        if gbot > gtop:
            for col in ("D", "E", "G"):
                _merge(ws, f"{col}{gtop}:{col}{gbot}")
        r = gbot + 1

    data_bot = r - 1

    # 媒體型態/地區/播出時段：延伸合併蓋住「延續列」(media_label 為空的補償/回饋列)，
    # 使左側區塊整齊為單一高格。萬家福表每列各有媒體名，不會被併。
    blk_top = None
    for rr, row in data_rows:
        if row["media_label"]:
            if blk_top is not None and rr - 1 > blk_top:
                for col in ("A", "B", "C"):
                    _merge(ws, f"{col}{blk_top}:{col}{rr - 1}")
            blk_top = rr
    if blk_top is not None and data_bot > blk_top:
        for col in ("A", "B", "C"):
            _merge(ws, f"{col}{blk_top}:{col}{data_bot}")
    # 素材提供時間（H）：整張表同一素材日，整塊合併蓋住所有資料列（含量販→樂家康）
    if data_bot > data_top:
        _merge(ws, f"H{data_top}:H{data_bot}")
    # 合計列
    ws.row_dimensions[r].height = 80.15
    _set(ws, f"B{r}", "合計", size=24)
    _merge(ws, f"B{r}:C{r}")
    _set(ws, f"D{r}", total_list, size=24, fmt=ACCT)
    _set(ws, f"F{r}", total_spots, size=24, fmt=SUM_FMT)
    _set(ws, f"G{r}", total_net, size=24, fmt=ACCT)
    for cidx, off in day_cols:
        _set(ws, f"{get_column_letter(cidx)}{r}", sum(s[off] for _, s in schedule_rows),
             size=24, bold=True, fmt=DAY_RED)
    total_row = r

    # 框線層次：整表外框 double、日欄縱線 hair、表頭下緣 double、合計上緣 double
    _box(ws, 8, 1, total_row, last_col, edge="double", inner="hair")
    for c in range(1, last_col + 1):
        _edge(ws, f"{get_column_letter(c)}10", bottom="double")   # 表頭區結束線
        _edge(ws, f"{get_column_letter(c)}{total_row}", top="double")  # 合計上緣
    # 主檔列 A 欄四邊 double
    for grp in _group_main_comp(sheet["rows"]):
        pass  # A 欄已在外框，主檔區塊視覺上已由 double 外框涵蓋

    # 費用區 F16:G19（四條細橫線的小表，無縱線）
    fr = total_row + 2
    f = sheet["fees"]
    fee_lines = [
        ("Budget (net)：", f["budget_net"]),
        (f"AC {int(f['ac_pct'])}%：", f["ac"]),
        ("5% Tax ：", f["tax"]),
        ("TOTAL ：", f["total"]),
    ]
    for i, (lab, val) in enumerate(fee_lines):
        rr = fr + i
        ws.row_dimensions[rr].height = 60
        _set(ws, f"F{rr}", lab, size=24, bold=True, align="right")
        _net_cell(ws, f"G{rr}", val, size=24, bold=True)
        _edge(ws, f"F{rr}", top="thin", bottom="thin")
        _edge(ws, f"G{rr}", top="thin", bottom="thin")

    # 備註 28pt 粗體
    rr = fr + len(fee_lines) + 1
    for line in model.get("remarks", []):
        ws.row_dimensions[rr].height = 55.75
        _set(ws, f"A{rr}", line, size=28, bold=True, align="left")
        _merge(ws, f"A{rr}:{get_column_letter(last_col)}{rr}")
        rr += 1
    return ws


# =============================================================================
# 佳聖 版型
# =============================================================================
def _render_ddrive(wb, sheet, model, made_date):
    days = (model["end_date"] - model["start_date"]).days + 1
    start_dt = model["start_date"]
    sec = sheet["seconds"]
    ws = wb.create_sheet(title=_sheet_name(model, sheet))
    _page(ws, margins=(0.0, 0.0, 0.4, 0.2))

    # 列印頁首/頁尾
    ws.oddHeader.center.text = "佳聖媒體  戶外媒體排期表"
    ws.oddFooter.left.text = ("          群主管:_______________部主管:_______________"
                              "課主管:_______________承辦PM:_______________")
    ws.oddFooter.right.text = "佳聖媒體: ____________________          "

    DAY0 = 9
    widths = {"A": 31.6, "B": 25.5, "C": 18.1, "D": 19.9, "E": 15.9, "F": 21.4, "G": 23.5, "H": 23.5}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for i in range(days):
        ws.column_dimensions[get_column_letter(DAY0 + i)].width = 8.5
    last_col = DAY0 + days - 1

    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 21.6
    for r in (5, 6, 7):
        ws.row_dimensions[r].height = 57
    for r in (9, 10, 11):
        ws.row_dimensions[r].height = 38

    _add_logo(ws, LOGO_DDRIVE, "A1", width=72, height=87)

    # 客戶/產品/刊期（標籤與值分開兩格；值前一空格；刊期短格式）
    _set(ws, "A5", "客戶：", size=22, bold=True, align="left")
    _set(ws, "B5", f" {model['client_name']}", size=22, bold=True, align="left")
    _set(ws, "A6", "產品：", size=22, bold=True, align="left")
    _set(ws, "B6", f" {model['product_name']}", size=22, bold=True, align="left")
    _set(ws, "A7", "刊期：", size=22, bold=True, align="left")
    _set(ws, "B7", f" {_period_short(model['start_date'], model['end_date'])}", size=22, bold=True, align="left")

    # 欄位表頭 9~11
    heads = [("A", "媒體"), ("B", "地區"), ("C", "託播秒數"), ("D", "播出時段"),
             ("E", "次數"), ("F", "素材\n提供時間"), ("G", "定價\n(Net Cost)"), ("H", "專案執行價\n(Net Cost)")]
    for col, txt in heads:
        _set(ws, f"{col}9", txt, size=18, bold=True, wrap=True)
        _merge(ws, f"{col}9:{col}11")
    _write_day_header(ws, DAY0, start_dt, days, 9, 10, 11, _cn_weekday, month_style="cn", size=18)

    day_cols = [(DAY0 + i, i) for i in range(days)]
    r = 12
    data_top = r
    schedule_rows = []
    total_spots = total_list = total_net = 0
    is_wjf = sheet["platform"] == ac.PLATFORM_WJF
    first_data = r
    for row in sheet["rows"]:
        ws.row_dimensions[r].height = 62
        _set(ws, f"A{r}", row["media_label"], size=18, wrap=True)
        _set(ws, f"B{r}", row["region_label"], size=18)
        _set(ws, f"C{r}", f"{sec}秒", size=18, wrap=True)
        _set(ws, f"D{r}", row["daypart"], size=18, wrap=True)
        _set(ws, f"E{r}", row["spots"], size=18)
        if row["material"]:
            _set(ws, f"F{r}", row["material"], size=18, fmt=MATERIAL_FMT)
        is_super = row["kind"] in (ac.KIND_SUPER, ac.KIND_SUPER_REBATE)
        gval = ac.NET_ON_MAG if is_super else row["list_total"]
        _net_cell(ws, f"G{r}", gval, size=18, fmt=NUM)
        _net_cell(ws, f"H{r}", row["net_display"], size=18, fmt=NUM, align="right")
        total_spots += row["spots"]
        if not is_super and isinstance(row["list_total"], (int, float)):
            total_list += row["list_total"]
        if isinstance(row["net_display"], (int, float)):
            total_net += row["net_display"]
        if row["schedule"] is None:
            c0, c1 = get_column_letter(DAY0), get_column_letter(last_col)
            _set(ws, f"{c0}{r}", row["spots"], size=18)
            if last_col > DAY0:
                _merge(ws, f"{c0}{r}:{c1}{r}")
        else:
            for cidx, off in day_cols:
                _set(ws, f"{get_column_letter(cidx)}{r}", row["schedule"][off], size=16, fmt=DAY_FMT)
            schedule_rows.append((r, row["schedule"]))
        r += 1
    data_bot = r - 1

    # 平台專屬合併
    if is_wjf:
        # 萬家福：C/F/H 合併量販+超市（前兩列）
        if data_bot >= first_data + 1:
            for col in ("C", "F", "H"):
                _merge(ws, f"{col}{first_data}:{col}{first_data + 1}")
            _net_cell(ws, f"H{first_data}", sheet["budget"] if not sheet["is_rebate_wave"] else ac.NET_REBATE,
                      size=18, fmt=NUM, align="right")
    else:
        # 全家：A/B/C/D/F 直向合併整塊
        if data_bot > first_data:
            for col in ("A", "B", "C", "D", "F"):
                _merge(ws, f"{col}{first_data}:{col}{data_bot}")

    # 小計列
    ws.row_dimensions[r].height = 41
    _set(ws, f"A{r}", "小計", size=18)
    _merge(ws, f"A{r}:D{r}")
    _set(ws, f"E{r}", total_spots, size=18, fmt="#,##0")
    _net_cell(ws, f"G{r}", total_list, size=18, fmt=NUM)
    # H = 實收（僅主檔）
    net_main = 0
    for row in sheet["rows"]:
        if row["kind"] == ac.KIND_MAIN and isinstance(row["net_display"], (int, float)):
            net_main += row["net_display"]
    _net_cell(ws, f"H{r}", net_main, size=18, fmt=NUM, align="right")
    for cidx, off in day_cols:
        _set(ws, f"{get_column_letter(cidx)}{r}", sum(s[off] for _, s in schedule_rows), size=16, fmt="#,##0")
    subtotal_row = r

    # 框線：外框 medium、內線 thin；小計底 medium
    _box(ws, 9, 1, subtotal_row, last_col, edge="medium", inner="thin")

    # 備註（左）＋ 請款；多月份請款每月一列往下排，避免長文字擠到右側費用欄
    br = subtotal_row + 2
    ws.row_dimensions[br].height = 52.7
    _set(ws, f"A{br}", "備註：", size=22, align="left")
    pn = model.get("payment_note", "")
    pay_parts = pn.split("、") if pn else [""]
    for idx, part in enumerate(pay_parts):
        rr = br + 1 + idx
        if idx > 0:
            ws.row_dimensions[rr].height = 40
            # 續月縮排＝前綴「* 請款金額：」寬度(≈6全形)，對齊首列「8月份」
            part = "　　　　　　" + part
        _set(ws, f"A{rr}", part, size=22, align="left", wrap=True)
        _merge(ws, f"A{rr}:E{rr}")

    # 費用框（F 標籤 / H 值，G 留空，無框線，無 $）
    f = sheet["fees"]
    fee_lines = [("Total Net Cost", f["net"]), ("VAT   (5%)", f["vat"]), ("Total Gross Cost", f["gross"])]
    for i, (lab, val) in enumerate(fee_lines):
        rr = subtotal_row + 2 + i
        _set(ws, f"F{rr}", lab, size=22, bold=True, align="left")
        _net_cell(ws, f"H{rr}", val, size=22, bold=True, fmt=NUM, align="right")
    return ws


# =============================================================================
# 凱絡版型
# =============================================================================
def _render_carat(wb, sheet, model, made_date):
    from datetime import timedelta
    days = (model["end_date"] - model["start_date"]).days + 1
    start_dt = model["start_date"]
    sec = sheet["seconds"]
    ws = wb.create_sheet(title=_sheet_name(model, sheet))
    _page(ws, margins=(0.0, 0.0, 0.3, 0.1))

    DAY0 = 11
    widths = {"A": 21.5, "B": 18.7, "C": 15.7, "D": 9.5, "E": 11.3, "F": 11.5,
              "G": 11.5, "H": 10.5, "I": 14.5, "J": 14.1}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for i in range(days):
        ws.column_dimensions[get_column_letter(DAY0 + i)].width = 8.6
    last_col = DAY0 + days - 1

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 30
    for r in (4, 5, 6, 7):
        ws.row_dimensions[r].height = 29.25

    made = made_date or model["start_date"]
    # 大標靠左、不合併
    _set(ws, "A1", "凱絡媒體服務(股)公司廣播媒體排期表", size=16, bold=True, align="left")
    # 右上抬頭
    _set(ws, "I2", "客   戶：", size=14, align="right")
    _set(ws, "J2", model["client_name"], size=14, align="left")
    _set(ws, "I3", "產   品：", size=14, align="right")
    _set(ws, "J3", model["product_name"], size=14, align="left")
    _set(ws, "I4", "日   期：", size=14, align="right")
    _set(ws, "J4", made.strftime("%Y/%m/%d"), size=14, align="left")
    _set(ws, "A4", f"{start_dt.year}年{start_dt.month}月", size=14, align="left")

    # 欄位表頭 5~7
    heads = [("A", "媒體別"), ("B", "地區"), ("C", "時段"), ("D", "素材"),
             ("E", "定價\n(檔/Net)"), ("F", "市場價\n(檔/Net)"), ("G", "統一價\n(檔/Net)"),
             ("H", "檔數"), ("I", "總價"), ("J", "專案價\n(Net)")]
    for col, txt in heads:
        _set(ws, f"{col}5", txt, size=12, bold=True, wrap=True)
        _merge(ws, f"{col}5:{col}7")
    _write_day_header(ws, DAY0, start_dt, days, 5, 6, 7, _en_weekday, month_style="en",
                      size=12, date_fmt="#,##0", weekend_fill_rows=(7,))
    # 最右加「總檔數」欄（對齊範本）
    tot_col = last_col + 1
    tl = get_column_letter(tot_col)
    ws.column_dimensions[tl].width = 9.5
    _set(ws, f"{tl}5", "總檔數", size=12, bold=True, wrap=True)
    _merge(ws, f"{tl}5:{tl}7")
    _box(ws, 5, 1, 7, tot_col, edge="medium", inner="thin")

    day_cols = [(DAY0 + i, i) for i in range(days)]
    r = 8
    data_top = r
    schedule_rows = []
    media_value = 0
    actual_net = sheet["fees"].get("subtotal") if isinstance(sheet["fees"].get("subtotal"), (int, float)) else 0

    # A 欄以（量販+超市）、（回饋量販+回饋超市）為組合併
    rows = sheet["rows"]
    groups = []
    i = 0
    while i < len(rows):
        if rows[i]["kind"] == ac.KIND_MAIN and i + 1 < len(rows) and rows[i + 1]["kind"] == ac.KIND_SUPER:
            groups.append([rows[i], rows[i + 1]]); i += 2
        elif rows[i]["kind"] == ac.KIND_REBATE and i + 1 < len(rows) and rows[i + 1]["kind"] == ac.KIND_SUPER_REBATE:
            groups.append([rows[i], rows[i + 1]]); i += 2
        else:
            groups.append([rows[i]]); i += 1

    for grp in groups:
        gtop = r
        for k, row in enumerate(grp):
            rr = r + k
            ws.row_dimensions[rr].height = 58
            _set(ws, f"B{rr}", row["region_label"], size=12, wrap=True)
            _set(ws, f"C{rr}", row["daypart"], size=12)
            _set(ws, f"D{rr}", f'{sec}"CM', size=12, fmt="@")
            _set(ws, f"E{rr}", row["list_per"], size=12, fmt=ACCT)
            _set(ws, f"F{rr}", row["market_per"], size=12, fmt=ACCT)
            _set(ws, f"G{rr}", row["uni_per"], size=12, fmt=ACCT)
            _set(ws, f"H{rr}", row["spots"], size=12, fmt=CARAT_H_FMT)
            _set(ws, f"I{rr}", row["uni_total"], size=12, fmt=ACCT)
            # 回饋列在表上顯示「聲活回饋」（內部 net_display 仍為專案回饋）
            jval = row["net_display"]
            if row["kind"] == ac.KIND_REBATE and jval == ac.NET_REBATE:
                jval = ac.CARAT_REBATE_LABEL
            _net_cell(ws, f"J{rr}", jval, size=12,
                      bold=(row["kind"] == ac.KIND_MAIN), align="center")
            media_value += row["uni_total"] if isinstance(row["uni_total"], (int, float)) else 0
            if row["schedule"] is None:
                c0, c1 = get_column_letter(DAY0), get_column_letter(last_col)
                _set(ws, f"{c0}{rr}", row["spots"], size=11)
                if last_col > DAY0:
                    _merge(ws, f"{c0}{rr}:{c1}{rr}")
            else:
                for cidx, off in day_cols:
                    _set(ws, f"{get_column_letter(cidx)}{rr}", row["schedule"][off], size=11, fmt=DAY_FMT)
                schedule_rows.append((rr, row["schedule"]))
            # 總檔數（最右欄）
            _set(ws, f"{tl}{rr}", row["spots"], size=12, fmt=CARAT_H_FMT)
        _set(ws, f"A{gtop}", grp[0]["media_label"], size=12, wrap=True)
        r += len(grp)
    data_bot = r - 1

    # 媒體別（A）整塊合併為單一格（全表同一媒體別）；
    # 全家表另把地區/時段/素材（B/C/D）併入回饋等延續列，使左側整齊。
    # 萬家福表量販/超市地區時段不同，故 B/C/D 逐列保留、只合併媒體別。
    is_wjf = sheet["platform"] == ac.PLATFORM_WJF
    if data_bot > data_top:
        _merge(ws, f"A{data_top}:A{data_bot}")
        if not is_wjf:
            for col in ("B", "C", "D"):
                _merge(ws, f"{col}{data_top}:{col}{data_bot}")

    _box(ws, data_top, 1, data_bot, tot_col, edge="medium", inner="thin")

    # 媒體總價值 / 優惠總價值（A 標籤、B 數字，medium 方框）
    mv_top = data_bot + 1
    _set(ws, f"A{mv_top}", "媒體總價值(NET)", size=12, bold=True, align="left")
    _set(ws, f"B{mv_top}", media_value, size=12, fmt=ACCT)
    _set(ws, f"A{mv_top + 1}", "優惠總價值(NET)", size=12, bold=True, align="left")
    _set(ws, f"B{mv_top + 1}", media_value - actual_net, size=12, fmt=ACCT)
    _box(ws, mv_top, 1, mv_top + 1, 2, edge="medium", inner="thin")

    # 費用區 I:J（純數字；只有 Grand-Total 上細線下雙線）
    f = sheet["fees"]
    ac_txt = "-" if f.get("ac_free") else f.get("ac", 0)
    fee_lines = [("Sub-Total", f["subtotal"]), ("A.C     3%", ac_txt),
                 ("VAT    5%", f["vat"]), ("Grand-Total", f["grand"])]
    for i, (lab, val) in enumerate(fee_lines):
        rr = data_bot + 1 + i
        _set(ws, f"I{rr}", lab, size=12, bold=True, align="left")
        _net_cell(ws, f"J{rr}", val, size=12, bold=True, fmt=NUM, align="right")
        if lab == "Grand-Total":
            _edge(ws, f"I{rr}", top="thin", bottom="double")
            _edge(ws, f"J{rr}", top="thin", bottom="double")

    # 簽核列（四個分開儲存格）
    sr = data_bot + 2 + len(fee_lines)
    _set(ws, f"A{sr}", "部主管：_________ ", size=12, align="left")
    _set(ws, f"D{sr}", "課主管：_________ ", size=12, align="left")
    _set(ws, f"I{sr}", "媒體窗口：_________", size=12, align="left")
    _set(ws, f"O{sr}", "承辦PM：_________", size=12, align="left")

    # 備註（A 欄「備 註」直向合併、medium 外框；內容 B 欄起）
    rmk = model.get("remarks", [])
    rk_top = sr + 1
    n = max(1, len(rmk))
    _set(ws, f"A{rk_top}", "備     註", size=12, bold=True, align="center", valign="center", wrap=True)
    if n > 1:
        _merge(ws, f"A{rk_top}:A{rk_top + n - 1}")
    for i, line in enumerate(rmk):
        rr = rk_top + i
        _set(ws, f"B{rr}", line, size=12, align="left", wrap=True)
        _merge(ws, f"B{rr}:{get_column_letter(last_col)}{rr}")
    _box(ws, rk_top, 1, rk_top + n - 1, last_col, edge="medium", inner=None)
    return ws


# =============================================================================
# 對外主函式
# =============================================================================
def _sheet_name(model, sheet):
    a = model["start_date"].strftime("%m%d")
    b = model["end_date"].strftime("%m%d")
    if model["agency"] == "2008傳媒":
        wan = round(sheet["budget"] / 10000) if sheet["budget"] else 0
        plat = "全家" if sheet["platform"] == ac.PLATFORM_FAMILY else "萬家福"
        name = f"{a}-{b}-{plat}-{wan}萬-{sheet['seconds']}秒"
    else:
        name = f"{sheet['platform']} {a}-{b}"
    for ch in ':\\/?*[]':
        name = name.replace(ch, "")
    return name[:31]


def generate_agency_excel(model, made_date=None):
    """依 model 產出代理商 Excel（每平台一個工作表），回傳 bytes。"""
    wb = Workbook()
    wb.remove(wb.active)
    agency = model["agency"]
    for sheet in model["sheets"]:
        if agency == "2008傳媒":
            _render_2008(wb, sheet, model, made_date)
        elif agency == "佳聖":
            _render_ddrive(wb, sheet, model, made_date)
        elif agency == "凱絡":
            _render_carat(wb, sheet, model, made_date)
        else:
            _render_2008(wb, sheet, model, made_date)
    if not wb.worksheets:
        wb.create_sheet(title="空")
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
